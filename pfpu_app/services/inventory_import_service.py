from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from ..config import DATA_DIR
from ..database import connect
from .backup_service import create_database_backup
from .excel_service import (
    clean,
    import_excel,
    to_float,
    to_int,
)


IMPORT_DIR = DATA_DIR / "imports"


def inventory_is_empty() -> bool:
    """
    Return True when no master inventory rows exist.
    """

    con = connect()

    try:
        count = con.execute(
            "SELECT COUNT(*) FROM item_master"
        ).fetchone()[0]

        return count == 0

    finally:
        con.close()


def stage_inventory_file(
    source_file,
    original_filename: str,
) -> dict:
    """
    Save an uploaded Excel file into PFPU's temporary
    inventory-import staging area.
    """

    filename = Path(
        original_filename or ""
    ).name

    if not filename.lower().endswith(".xlsx"):
        return {
            "success": False,
            "message": (
                "Inventory import must be an .xlsx Excel file."
            ),
        }

    IMPORT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    token = uuid4().hex

    staged_path = (
        IMPORT_DIR
        / f"{token}.xlsx"
    )

    try:
        source_file.seek(0)

        with staged_path.open("wb") as output:
            while True:
                chunk = source_file.read(
                    1024 * 1024
                )

                if not chunk:
                    break

                output.write(chunk)

    except Exception as exc:

        if staged_path.exists():
            staged_path.unlink()

        return {
            "success": False,
            "message": (
                "Unable to save uploaded inventory file: "
                f"{exc}"
            ),
        }

    return {
        "success": True,
        "message": "Inventory file uploaded.",
        "token": token,
        "path": staged_path,
        "filename": filename,
    }


def get_staged_import_path(
    token: str,
) -> Path | None:
    """
    Resolve a safe staged import token.
    """

    token = (
        token or ""
    ).strip().lower()

    if (
        not token
        or len(token) != 32
        or not all(
            character in "0123456789abcdef"
            for character in token
        )
    ):
        return None

    path = (
        IMPORT_DIR
        / f"{token}.xlsx"
    )

    if not path.exists():
        return None

    return path


def analyze_inventory_file(
    path: Path,
) -> dict:
    """
    Inspect an uploaded workbook without changing the database.
    """

    path = Path(path)

    if not path.exists():
        return {
            "success": False,
            "message": "Uploaded inventory file was not found.",
        }

    try:
        workbook = load_workbook(
            path,
            data_only=True,
            read_only=True,
        )

    except Exception as exc:
        return {
            "success": False,
            "message": (
                "Unable to read the Excel workbook: "
                f"{exc}"
            ),
        }

    preview = []
    warnings = []
    valid_rows = 0

    try:

        if "Equipment Mapping" in workbook.sheetnames:

            format_name = "Equipment Mapping"
            worksheet = workbook["Equipment Mapping"]

            for row in worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ):

                if (
                    not row
                    or len(row) < 5
                    or not row[4]
                ):
                    continue

                values = list(row) + [None] * 8

                original_row = values[0]
                category = values[1]
                prefix = values[2]
                qty = values[3]
                description = values[4]
                unit_value = values[5]
                total_value = values[6]
                priority = values[7]

                description = clean(
                    description
                )

                if not description:
                    continue

                valid_rows += 1

                if len(preview) < 12:
                    preview.append(
                        {
                            "original_row": to_int(
                                original_row
                            ),
                            "category": (
                                clean(category)
                                or "Uncategorized"
                            ),
                            "prefix": (
                                clean(prefix)
                                or "AST"
                            ),
                            "qty": to_int(qty),
                            "description": description,
                            "unit_value": to_float(
                                unit_value
                            ),
                            "total_value": to_float(
                                total_value
                            ),
                            "priority": clean(
                                priority
                            ),
                        }
                    )

        elif "EQUIPMENT" in workbook.sheetnames:

            format_name = "EQUIPMENT"
            worksheet = workbook["EQUIPMENT"]

            for row_number in range(
                6,
                worksheet.max_row + 1,
            ):

                qty = to_int(
                    worksheet.cell(
                        row_number,
                        1,
                    ).value
                )

                description = clean(
                    worksheet.cell(
                        row_number,
                        2,
                    ).value
                )

                if not qty or not description:
                    continue

                valid_rows += 1

                if len(preview) < 12:
                    preview.append(
                        {
                            "original_row": row_number,
                            "category": "Imported",
                            "prefix": "AST",
                            "qty": qty,
                            "description": description,
                            "unit_value": to_float(
                                worksheet.cell(
                                    row_number,
                                    4,
                                ).value
                            ),
                            "total_value": to_float(
                                worksheet.cell(
                                    row_number,
                                    5,
                                ).value
                            ),
                            "priority": "Needs review",
                        }
                    )

            warnings.append(
                "The EQUIPMENT format does not provide "
                "PFPU categories, prefixes, or tracking "
                "priority. Those fields will use imported "
                "defaults and can be reviewed afterward."
            )

        else:

            return {
                "success": False,
                "message": (
                    "PFPU could not recognize this workbook. "
                    "Expected a sheet named 'Equipment Mapping' "
                    "or 'EQUIPMENT'."
                ),
            }

    finally:
        workbook.close()

    if valid_rows == 0:
        return {
            "success": False,
            "message": (
                "The workbook was recognized but no valid "
                "inventory rows were found."
            ),
        }

    return {
        "success": True,
        "message": (
            f"{valid_rows} inventory rows are ready to import."
        ),
        "format": format_name,
        "row_count": valid_rows,
        "preview": preview,
        "warnings": warnings,
    }


def confirm_inventory_import(
    token: str,
) -> dict:
    """
    Import a previously validated staged workbook.

    The first-release importer intentionally requires an empty
    master inventory to prevent accidental duplicate imports.
    """

    path = get_staged_import_path(
        token
    )

    if path is None:
        return {
            "success": False,
            "message": (
                "The staged inventory import could not be found."
            ),
        }

    analysis = analyze_inventory_file(
        path
    )

    if not analysis["success"]:
        return analysis

    if not inventory_is_empty():
        return {
            "success": False,
            "message": (
                "Inventory already contains items. "
                "Initial Excel import is only available "
                "for an empty inventory database."
            ),
        }

    backup = create_database_backup(
        "before-inventory-import"
    )

    if not backup["success"]:
        return {
            "success": False,
            "message": (
                "Inventory import was stopped because PFPU "
                "could not create its safety backup. "
                + backup["message"]
            ),
        }

    try:
        imported = import_excel(
            path
        )

    except Exception as exc:
        return {
            "success": False,
            "message": (
                "Inventory import failed: "
                f"{exc}"
            ),
        }

    if imported <= 0:
        return {
            "success": False,
            "message": (
                "No inventory rows were imported."
            ),
        }

    try:
        path.unlink()
    except OSError:
        pass

    return {
        "success": True,
        "message": (
            f"Inventory import complete. "
            f"{imported} inventory rows were added."
        ),
        "imported": imported,
    }