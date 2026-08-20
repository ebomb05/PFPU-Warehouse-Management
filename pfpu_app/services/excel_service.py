from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ..config import DATA_DIR
from ..database import connect


def clean(value):
    return "" if value is None else str(value).strip()


def to_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def to_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def import_excel(path: Path):
    wb = load_workbook(path, data_only=True)
    con = connect()
    cur = con.cursor()
    imported = 0

    if "Equipment Mapping" in wb.sheetnames:
        ws = wb["Equipment Mapping"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5 or not row[4]:
                continue

            original_row, category, prefix, qty, desc, unit_value, total_value, priority = row[:8]
            cur.execute(
                """
                INSERT INTO item_master(
                    original_row, category, prefix, qty_total, description,
                    unit_value, total_value, tracking_priority
                )
                VALUES(?,?,?,?,?,?,?,?)
                """,
                (
                    to_int(original_row),
                    clean(category) or "Uncategorized",
                    clean(prefix) or "AST",
                    to_int(qty),
                    clean(desc),
                    to_float(unit_value),
                    to_float(total_value),
                    clean(priority),
                ),
            )
            imported += 1

    elif "EQUIPMENT" in wb.sheetnames:
        ws = wb["EQUIPMENT"]
        for row_number in range(6, ws.max_row + 1):
            qty = to_int(ws.cell(row_number, 1).value)
            desc = clean(ws.cell(row_number, 2).value)
            if qty and desc:
                cur.execute(
                    """
                    INSERT INTO item_master(
                        original_row, category, prefix, qty_total, description,
                        unit_value, total_value, tracking_priority
                    )
                    VALUES(?,?,?,?,?,?,?,?)
                    """,
                    (
                        row_number,
                        "Imported",
                        "AST",
                        qty,
                        desc,
                        to_float(ws.cell(row_number, 4).value),
                        to_float(ws.cell(row_number, 5).value),
                        "Needs review",
                    ),
                )
                imported += 1

    con.commit()
    con.close()
    return imported


def export_excel_file():
    con = connect()
    wb = Workbook()

    exports = [
        ("Inventory Master", "SELECT * FROM item_master"),
        ("Assets", "SELECT * FROM assets"),
        ("Jobs", "SELECT * FROM jobs"),
        ("Job Lines", "SELECT * FROM job_lines"),
        ("Scan Log", "SELECT * FROM scan_log"),
    ]

    for index, (sheet_name, query) in enumerate(exports):
        ws = wb.active if index == 0 else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        cur = con.execute(query)
        headers = [column[0] for column in cur.description]
        ws.append(headers)

        for row in cur.fetchall():
            ws.append([row[header] for header in headers])

        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = min(
                max(len(str(column[0].value or "")) + 2, 12),
                40,
            )

    output = DATA_DIR / f"PFPU_inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output)
    con.close()
    return output
